# !!! make sure u close all the excel files before running this !!!

import pandas as pd
import win32com.client as win32
import glob
import os
import shutil
import tempfile
import time
import traceback
import sys
from pathlib import Path

import config
import private
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

folder_path = private.PATH
output_file = 'MasterReport.xlsx'

def import_original_sheets(master_path, source_folder):
    if not win32:
        return
    print(f"{config.colors.BOLD}{config.colors.YELLOW}[STATUS]{config.colors.RESET} Importing original sheets")
    
    master_path = os.path.abspath(master_path)
    source_folder = os.path.abspath(source_folder)
    if not os.path.exists(master_path):
        print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} Master file not found at {master_path}")
        return
    
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    
    try:
        wb_master = excel.Workbooks.Open(master_path)
        
        # grab all the excel files
        all_files = glob.glob(os.path.join(source_folder, "**", "*.xlsx"), recursive=True)
        
        for filename in all_files:
            # skip master file
            if os.path.abspath(filename) == master_path:
                continue
            # skip excel temp lock files
            if os.path.basename(filename).startswith('~$'):
                continue
            
            temp_dir = tempfile.gettempdir()
            temp_filename = os.path.join(temp_dir, f"temp_{int(time.time())}_{os.path.basename(filename)}")
            
            try:
                # copy to temp to bcs long path issues happen bcs these ppl wanna name their folders so long for some reason
                shutil.copy2(filename, temp_filename)
                
                wb_source = excel.Workbooks.Open(temp_filename)
                
                # copy the excel sheets to master
                for i, sheet in enumerate(wb_source.Sheets):
                    original_sheet_count = wb_master.Sheets.Count
                    sheet.Copy(None, wb_master.Sheets(original_sheet_count))
                    time.sleep(0.5) # stability pause
                    
                    if wb_master.Sheets.Count > original_sheet_count:
                        print(f"{config.colors.BOLD}{config.colors.YELLOW}[STATUS]{config.colors.RESET} Importing sheet from {os.path.basename(filename)}")
                        try:
                            # remove extension
                            base_name = os.path.splitext(os.path.basename(filename))[0]
                            # if theres multiple sheets, we'll append the index to avoid name collision
                            if wb_source.Sheets.Count > 1:
                                new_name = f"{base_name[:28]}_{i+1}"
                            else:
                                new_name = base_name[:31] # excel name limit is 31
                            
                            # check for name collision
                            existing_names = [wb_master.Sheets(k).Name for k in range(1, wb_master.Sheets.Count)]
                            
                            original_new_name = new_name
                            collision_count = 1
                            while new_name in existing_names:
                                suffix = f"_{collision_count}"
                                max_len = 31 - len(suffix)
                                new_name = f"{original_new_name[:max_len]}{suffix}"
                                collision_count += 1
                                
                            wb_master.Sheets(wb_master.Sheets.Count).Name = new_name
                            print(f"{config.colors.BOLD}{config.colors.GREEN}[SUCCESS]{config.colors.RESET} Finished sheet importing from {os.path.basename(filename)}")
                        except Exception as error:
                            print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} Could not rename sheet: {error}")
                    else:
                        print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} Sheet copy failed for {os.path.basename(filename)}")  
                wb_source.Close(SaveChanges=False)
            except Exception as error:
                print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} Error processing {filename}: {error}")
            finally:
                # cleanup temp file
                if os.path.exists(temp_filename):
                    try:
                        os.remove(temp_filename)
                    except:
                        pass
        
        wb_master.Save()
        wb_master.Close()
    except Exception as error:
        print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} Could not import excel sheets: {error}")
    finally:
        try:
            excel.Quit()
        except:
            pass

def merge_excel_sheets(path, output_name):
    all_files = glob.glob(os.path.join(path, "**", "*.xlsx"), recursive=True)
    print(f"{config.colors.BOLD}{config.colors.YELLOW}[STATUS]{config.colors.RESET} {len(all_files)} excel files to process")
    all_data = [] # hold tables from each file
    
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    
    try:
        for filename in all_files:
            if os.path.basename(filename) == output_name:
                continue
            # skip excel temp lock files
            if os.path.basename(filename).startswith('~$'):
                continue
            
            temp_dir = tempfile.gettempdir()
            temp_filename = os.path.join(temp_dir, f"temp_{int(time.time())}_{os.path.basename(filename)}")
            
            row_data = {}
            try:
                # copy to temp for long path issues
                shutil.copy2(filename, temp_filename)
                
                wb = excel.Workbooks.Open(temp_filename, ReadOnly=True, UpdateLinks=False)
                sheet = wb.ActiveSheet
                row_data['Source File'] = os.path.basename(filename)
                
                # extract cells
                for col_name, cell_address in config.extraction_cells.items():
                    try:
                        # check for list
                        if isinstance(cell_address, list):
                            values = []
                            for addr in cell_address:
                                val = sheet.Range(addr).Value
                                if val:
                                    values.append(str(val))
                            row_data[col_name] = " ".join(values) if values else None
                        else:
                            val = sheet.Range(cell_address).Value
                            row_data[col_name] = val
                    except Exception as error:
                        row_data[col_name] = None
                        print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} Could not read cell {cell_address} in {filename}")
                all_data.append(row_data)
                print(f"{config.colors.BOLD}{config.colors.YELLOW}[STATUS]{config.colors.RESET} Read {os.path.basename(filename)}")
                wb.Close(SaveChanges=False)
            except Exception as error:
                print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} Could not read {filename}: {error}")
            finally:
                if os.path.exists(temp_filename):
                    try:
                        os.remove(temp_filename)
                    except:
                        pass
    finally:
        excel.Quit()
        
    # combine tables into one
    if all_data:
        combined_df = pd.DataFrame(all_data)
        
        # export to new excel file
        output_path = os.path.join(path, output_name)
        
        # format output
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            combined_df.to_excel(writer, index=False, sheet_name='Master Report')
            worksheet = writer.sheets['Master Report']
            
            # auto calc column width
            for i, column in enumerate(combined_df.columns):
                max_len = len(str(column))
                if not combined_df.empty:
                    data_len = combined_df[column].astype(str).map(len).max()
                    if not pd.isna(data_len):
                        max_len = max(max_len, data_len)
                
                column_letter = get_column_letter(i + 1)
                
                # formatting requirements
                is_source_file = column == "Source File"
                should_center = column in config.centered_columns
                should_wrap = max_len + 2 > 50
                should_green = column in getattr(config, 'green_columns', [])
                should_red = column in getattr(config, 'red_columns', []) # btw this only turns red if there is text present in the cell (some days theres no event)
                
                # column width wrapping
                if should_wrap:
                    worksheet.column_dimensions[column_letter].width = 50
                else:
                    worksheet.column_dimensions[column_letter].width = max_len + 2
                    
                # green highlighting (for income col)
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                # red highlighting (for daily events)
                red_fill = PatternFill(start_color="FFC1C1", end_color="FFC1C1", fill_type="solid")
                    
                # apply
                for cell in worksheet[column_letter]:
                    # default
                    horiz = None
                    vert = None
                    
                    
                    if should_center:
                        horiz = 'center'
                        vert = 'center'
                    elif is_source_file:
                        vert = 'center'
                        
                    if should_wrap:
                        if not vert: vert = 'center'
                        
                    if should_center or should_wrap or is_source_file:
                        cell.alignment = Alignment(horizontal=horiz, vertical=vert, wrap_text=should_wrap)
                
                    if should_green:
                        cell.fill = green_fill
                        
                    if should_red and cell.value: # cell.value checks for text inside the cell
                        cell.fill = red_fill
                
        print(f"{config.colors.BOLD}{config.colors.GREEN}[SUCCESS]{config.colors.RESET} All files merged into: {output_path}")
    else:
        print(f"{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} No data found to merge")

if __name__ == "__main__":
    try:
        merge_excel_sheets(folder_path, output_file)
        import_original_sheets(os.path.join(folder_path, output_file), folder_path)
    except Exception as error:
        print(f"\n{config.colors.BOLD}{config.colors.RED}[ERROR]{config.colors.RESET} {error}")
        traceback.print_exc()
    
    print(f"\n{config.colors.BOLD}{config.colors.GREEN}Done.{config.colors.RESET}")
    input(f"{config.colors.CYAN}Press Enter to exit.{config.colors.RESET}")